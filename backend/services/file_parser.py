"""Generic file parser service supporting PDF and Excel formats."""

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class FileParser:
    """Parses uploaded financial documents.

    - Excel: uses openpyxl to read standard financial statement columns.
    - PDF: uses pypdf to extract text then regex-match key financial terms.
    """

    # ── Column-keyword mappings ──────────────────────────────────────────

    _FINANCIAL_KEYWORDS: dict[str, list[str]] = {
        "total_assets": ["资产总计", "总资产", "total_assets", "total assets"],
        "total_liabilities": ["负债合计", "总负债", "total_liabilities", "total liabilities"],
        "revenue": ["营业收入", "营收", "收入", "revenue", "total revenue"],
        "net_profit": ["净利润", "净利", "net_profit", "net profit"],
        "operating_cash_flow": [
            "经营活动现金流",
            "经营现金流",
            "经营活动产生的现金流量净额",
            "operating_cash_flow",
            "operating cash flow",
        ],
    }

    _BANK_STMT_KEYWORDS: dict[str, list[str]] = {
        "date": ["交易日期", "日期", "date", "交易时间"],
        "summary": ["摘要", "摘要信息", "备注", "summary", "description"],
        "inflow": ["借方", "收入", "存入", "贷方收入", "inflow", "debit"],
        "outflow": ["贷方", "支出", "付出", "借方支出", "outflow", "credit"],
        "balance": ["余额", "当前余额", "balance"],
    }

    # ── Public API ───────────────────────────────────────────────────────

    async def parse_financial_report(self, file_path: str, file_type: str) -> dict[str, Any]:
        """Extract key financial indicators from an uploaded report file.

        Returns a dict with keys:
            total_assets, total_liabilities, revenue, net_profit,
            operating_cash_flow, parsed_tables
        """
        ext = Path(file_path).suffix.lower()
        if ext == ".xlsx":
            return await self._parse_excel_financial(file_path)
        if ext == ".pdf":
            return await self._parse_pdf_financial(file_path)
        raise ValueError(f"Unsupported file type for financial report: {ext}")

    async def parse_bank_statement(self, file_path: str, file_type: str) -> dict[str, Any]:
        """Extract transaction statistics from a bank statement file.

        Returns a dict with keys:
            total_inflow, total_outflow, avg_daily_balance, ending_balance,
            transaction_count, transaction_summary, anomaly_flags
        """
        ext = Path(file_path).suffix.lower()
        if ext == ".xlsx":
            return await self._parse_excel_bank_statement(file_path)
        if ext == ".pdf":
            return await self._parse_pdf_bank_statement(file_path)
        # Image formats
        if ext in {".jpg", ".png"}:
            return {
                "total_inflow": None,
                "total_outflow": None,
                "avg_daily_balance": None,
                "ending_balance": None,
                "transaction_count": 0,
                "transaction_summary": {},
                "anomaly_flags": {"image_file": "需要人工审核"},
            }
        raise ValueError(f"Unsupported file type for bank statement: {ext}")

    async def parse_credit_report(self, file_path: str, file_type: str) -> dict[str, Any]:
        """MVP: just note that parsing is not yet fully implemented."""
        return {"parse_note": "MVP阶段仅提供文件存储，解析结果待完善"}

    # ── Excel parsers ────────────────────────────────────────────────────

    @staticmethod
    def _find_column(headers: list[str], keywords: list[str]) -> int | None:
        """Return the index of the first header matching any keyword."""
        for idx, hdr in enumerate(headers):
            if hdr is None:
                continue
            normalized = str(hdr).strip().lower()
            for kw in keywords:
                if kw.lower() in normalized:
                    return idx
        return None

    @classmethod
    def _parse_number(cls, value: Any) -> float | None:
        """Try to convert a cell value to float."""
        if value is None:
            return None
        try:
            s = str(value).replace(",", "").replace(" ", "").strip()
            if s in ("", "-", "—"):
                return None
            return float(s)
        except (ValueError, TypeError):
            return None

    async def _parse_excel_financial(self, file_path: str) -> dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            return self._fallback_result(
                "openpyxl 未安装，无法解析 Excel",
            )

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active

        # Try to find header row (scan first 10 rows)
        headers: list[str] = []
        header_row_idx: int | None = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            row_texts = [str(cell.value).strip() if cell.value else "" for cell in row]
            if any(kw in " ".join(row_texts) for kw in ["资产", "负债", "收入", "利润"]):
                headers = row_texts
                header_row_idx = row_idx
                break

        result: dict[str, Any] = {
            "total_assets": None,
            "total_liabilities": None,
            "revenue": None,
            "net_profit": None,
            "operating_cash_flow": None,
            "parsed_tables": [],
        }

        if not headers:
            wb.close()
            return self._fallback_result("未在 Excel 中找到匹配的表头行")

        # Map keyword groups to column indices
        col_map: dict[str, int] = {}
        for field, keywords in self._FINANCIAL_KEYWORDS.items():
            idx = self._find_column(headers, keywords)
            if idx is not None:
                col_map[field] = idx

        # Scan data rows for numeric values in matched columns
        values_found: dict[str, list[float]] = {f: [] for f in col_map}
        if header_row_idx:
            for row in ws.iter_rows(
                min_row=header_row_idx + 1,
                values_only=True,
            ):
                for field, col_idx in col_map.items():
                    if col_idx < len(row):
                        val = self._parse_number(row[col_idx])
                        if val is not None:
                            values_found[field].append(val)

        wb.close()

        # Pick the last (or max) value for each field
        for field in col_map:
            vals = values_found[field]
            if vals:
                result[field] = round(vals[-1], 2)

        result["parsed_tables"].append({
            "headers": headers[: min(len(headers), 20)],
            "matched_columns": dict(col_map),
        })
        return result

    async def _parse_excel_bank_statement(self, file_path: str) -> dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            return {
                "total_inflow": None,
                "total_outflow": None,
                "avg_daily_balance": None,
                "ending_balance": None,
                "transaction_count": 0,
                "transaction_summary": {},
                "anomaly_flags": {"parse_error": "openpyxl 未安装"},
            }

        from datetime import date as date_type

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active

        # Find header row
        headers: list[str] = []
        header_row_idx: int | None = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            row_texts = [str(cell.value).strip() if cell.value else "" for cell in row]
            if any(kw in " ".join(row_texts) for kw in ["日期", "摘要", "借方", "贷方", "余额", "inflow", "outflow"]):
                headers = row_texts
                header_row_idx = row_idx
                break

        col_map: dict[str, int] = {}
        if headers:
            for field, keywords in self._BANK_STMT_KEYWORDS.items():
                idx = self._find_column(headers, keywords)
                if idx is not None:
                    col_map[field] = idx

        total_inflow = 0.0
        total_outflow = 0.0
        balances: list[float] = []
        ending_balance: float | None = None
        transaction_count = 0
        monthly_inflow: dict[str, float] = {}
        monthly_outflow: dict[str, float] = {}
        anomaly_flags: dict[str, Any] = {}

        if header_row_idx:
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                inflow_val = self._parse_number(row[col_map["inflow"]]) if "inflow" in col_map and col_map["inflow"] < len(row) else None
                outflow_val = self._parse_number(row[col_map["outflow"]]) if "outflow" in col_map and col_map["outflow"] < len(row) else None

                if inflow_val is not None and inflow_val > 0:
                    total_inflow += inflow_val
                    transaction_count += 1
                    # Monthly bucket
                    if "date" in col_map and col_map["date"] < len(row):
                        month_key = str(row[col_map["date"]])[:7]
                        monthly_inflow[month_key] = monthly_inflow.get(month_key, 0) + inflow_val
                if outflow_val is not None and outflow_val > 0:
                    total_outflow += outflow_val
                    if "date" in col_map and col_map["date"] < len(row):
                        month_key = str(row[col_map["date"]])[:7]
                        monthly_outflow[month_key] = monthly_outflow.get(month_key, 0) + outflow_val

                balance_val = self._parse_number(row[col_map["balance"]]) if "balance" in col_map and col_map["balance"] < len(row) else None
                if balance_val is not None:
                    balances.append(balance_val)
                    ending_balance = balance_val

        avg_daily_balance = round(sum(balances) / len(balances), 2) if balances else None

        # Anomaly detection (basic)
        if total_inflow > 0 and total_outflow > 0:
            ratio = total_inflow / total_outflow
            if ratio > 5 or ratio < 0.2:
                anomaly_flags["extreme_inflow_outflow_ratio"] = round(ratio, 2)

        wb.close()

        return {
            "total_inflow": round(total_inflow, 2),
            "total_outflow": round(total_outflow, 2),
            "avg_daily_balance": avg_daily_balance,
            "ending_balance": round(ending_balance, 2) if ending_balance else None,
            "transaction_count": transaction_count,
            "transaction_summary": {
                "monthly_inflow": monthly_inflow,
                "monthly_outflow": monthly_outflow,
            },
            "anomaly_flags": anomaly_flags,
        }

    # ── PDF parsers ──────────────────────────────────────────────────────

    async def _parse_pdf_financial(self, file_path: str) -> dict[str, Any]:
        try:
            from pypdf import PdfReader
        except ImportError:
            return self._fallback_result("pypdf 未安装，无法解析 PDF")

        reader = PdfReader(file_path)
        full_text = "\n".join(page.extract_text() or "" for page in reader.pages)

        result: dict[str, Any] = {
            "total_assets": None,
            "total_liabilities": None,
            "revenue": None,
            "net_profit": None,
            "operating_cash_flow": None,
            "parsed_tables": [{"text_snippet": full_text[:500]}],
        }

        patterns: dict[str, re.Pattern] = {
            "total_assets": re.compile(r"资产总计\s*[:：]?\s*([\d,\.]+)"),
            "total_liabilities": re.compile(r"负债合计\s*[:：]?\s*([\d,\.]+)"),
            "revenue": re.compile(r"营业收入?\s*[:：]?\s*([\d,\.]+)"),
            "net_profit": re.compile(r"净利润\s*[:：]?\s*([\d,\.]+)"),
            "operating_cash_flow": re.compile(r"经营活动现金流量净额?\s*[:：]?\s*([\d,\.]+)"),
        }

        for field, pattern in patterns.items():
            match = pattern.search(full_text)
            if match:
                val_str = match.group(1).replace(",", "")
                try:
                    result[field] = round(float(val_str), 2)
                except ValueError:
                    pass

        return result

    async def _parse_pdf_bank_statement(self, file_path: str) -> dict[str, Any]:
        """MVP: extract text from PDF bank statement, return minimal stats."""
        try:
            from pypdf import PdfReader
        except ImportError:
            return {
                "total_inflow": None,
                "total_outflow": None,
                "avg_daily_balance": None,
                "ending_balance": None,
                "transaction_count": 0,
                "transaction_summary": {},
                "anomaly_flags": {"parse_error": "pypdf 未安装"},
            }

        reader = PdfReader(file_path)
        full_text = "\n".join(page.extract_text() or "" for page in reader.pages)

        # Try to find any numbers that look like amounts
        numbers = re.findall(r"[\d,]+\.?\d*", full_text)
        float_numbers = []
        for n in numbers:
            try:
                float_numbers.append(float(n.replace(",", "")))
            except ValueError:
                pass

        return {
            "total_inflow": None,
            "total_outflow": None,
            "avg_daily_balance": None,
            "ending_balance": None,
            "transaction_count": len(float_numbers),
            "transaction_summary": {"pdf_text_length": len(full_text)},
            "anomaly_flags": {"note": "PDF 银行流水暂不支持自动解析，需要人工审核"},
        }

    # ── Helpers ──────────────────────────────────────────────────────────

    @staticmethod
    def _fallback_result(reason: str) -> dict[str, Any]:
        return {
            "total_assets": None,
            "total_liabilities": None,
            "revenue": None,
            "net_profit": None,
            "operating_cash_flow": None,
            "parsed_tables": [],
            "_parse_error": reason,
        }
